import io
import re
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill

from .address_enrichment import enrich_address_fields, extract_cep_from_text

ADDRESS_COLUMNS = {
    "endereco": (
        "ENDERECO",
        "ENDEREÇO",
        "ENDereco",
        "RUA",
        "LOGRADOURO",
        "ADDRESS",
        "ENDEREÇO:",
    ),
    "numero": ("NUMERO", "NÚMERO", "NRO", "NUM"),
    "bairro": ("BAIRRO",),
    "cep": ("CEP.1", "CEP", "CEP1"),
    "cidade": ("CIDADE", "MUNICIPIO", "MUNICÍPIO"),
    "uf": ("UF", "ESTADO"),
}

FILL_DENTRO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PROXIMO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FONT_DESTAQUE = Font(bold=True)

MANCHA_VERDE_VIABILITY = "Dentro da mancha"


def filter_mancha_verde_results(resultados: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        item
        for item in resultados
        if item.get("viabilidade") == MANCHA_VERDE_VIABILITY and item.get("status") == "ok"
    ]


def _normalize_column_name(name: str) -> str:
    text = str(name).strip().strip('"').strip("'")
    text = re.sub(r"\s*\[#.*?\]", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip().rstrip(":").strip()
    return text.upper()


def _find_column(normalized: dict[str, str], candidates: tuple[str, ...]) -> str | None:
    for candidate in candidates:
        key = _normalize_column_name(candidate)
        if key in normalized:
            return normalized[key]

    for candidate in candidates:
        key = _normalize_column_name(candidate)
        for norm_name, original in normalized.items():
            if norm_name.startswith(key) or norm_name.endswith(key):
                return original

    return None


def _build_column_map(columns: list[str]) -> dict[str, str | None]:
    normalized = {_normalize_column_name(col): col for col in columns}
    mapping: dict[str, str | None] = {}

    for field, candidates in ADDRESS_COLUMNS.items():
        mapping[field] = _find_column(normalized, candidates)

    if not mapping.get("endereco"):
        for norm_name, original in normalized.items():
            if any(token in norm_name for token in ("ENDERECO", "ENDEREÇO", "ADDRESS", "LOGRADOURO")):
                mapping["endereco"] = original
                break

    if mapping.get("endereco") and not mapping.get("cidade"):
        mapping["cidade"] = _find_column(normalized, ADDRESS_COLUMNS["cidade"])
    if mapping.get("endereco") and not mapping.get("uf"):
        mapping["uf"] = _find_column(normalized, ADDRESS_COLUMNS["uf"])
    if mapping.get("endereco") and not mapping.get("cep"):
        mapping["cep"] = _find_column(normalized, ADDRESS_COLUMNS["cep"])

    return mapping


def _clean_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", ".", "cep"}:
        return ""
    return text


def _format_cep(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        digits = str(int(value))
    else:
        digits = re.sub(r"\D", "", str(value))

    if not digits:
        return ""

    digits = digits.zfill(8)[:8]
    return f"{digits[:5]}-{digits[5:]}"


def _format_numero(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            return str(int(value))
        return str(value)

    return str(value).strip()


def _looks_like_full_address(text: str) -> bool:
    if len(text) < 35:
        return False
    upper = text.upper()
    if re.search(r"\d{5}-?\d{3}", text):
        return True
    if text.count(" - ") >= 2 and re.search(r"\b[A-Z]{2}\b", upper):
        return True
    return False


def to_google_search_query(address: str) -> str:
    text = re.sub(r"\s+", " ", address.strip())
    text = re.sub(r"\bCEP\s*", "", text, flags=re.IGNORECASE)
    parts = [part.strip() for part in text.split(" - ") if part.strip()]
    query = ", ".join(parts)
    if "brasil" not in query.lower() and "brazil" not in query.lower():
        query += ", Brasil"
    return query


def _format_address(row: pd.Series, column_map: dict[str, str | None]) -> tuple[str | None, bool]:
    endereco = _clean_text(row.get(column_map["endereco"])) if column_map["endereco"] else ""
    numero = _format_numero(row.get(column_map["numero"])) if column_map["numero"] else ""
    bairro = _clean_text(row.get(column_map["bairro"])) if column_map["bairro"] else ""
    cep = _format_cep(row.get(column_map["cep"])) if column_map["cep"] else ""
    cidade = _clean_text(row.get(column_map["cidade"])) if column_map["cidade"] else ""
    uf = _clean_text(row.get(column_map["uf"])) if column_map["uf"] else ""

    if not cep:
        cep = extract_cep_from_text(endereco)

    if endereco and _looks_like_full_address(endereco):
        return to_google_search_query(endereco), False

    enriched_fields, enriched = enrich_address_fields(
        endereco=endereco,
        bairro=bairro,
        cep=cep,
        cidade=cidade,
        uf=uf,
    )
    endereco = enriched_fields["endereco"]
    bairro = enriched_fields["bairro"]
    cep = enriched_fields["cep"]
    cidade = enriched_fields["cidade"]
    uf = enriched_fields["uf"]

    parts: list[str] = []

    if endereco:
        street = endereco
        if numero and numero not in endereco:
            street = f"{street}, {numero}"
        parts.append(street)
    elif bairro:
        parts.append(bairro)

    if bairro and endereco and bairro.upper() not in endereco.upper():
        parts.append(bairro)

    if cep and cep.replace("-", "") not in endereco.replace("-", ""):
        parts.append(f"CEP {cep}")

    if cidade:
        city_part = cidade
        if uf:
            city_part = f"{city_part} - {uf}"
        if cidade.upper() not in endereco.upper():
            parts.append(city_part)
    elif uf and uf.upper() not in endereco.upper():
        parts.append(uf)

    if not parts:
        return None, enriched

    return to_google_search_query(" - ".join(parts)), enriched


def _read_csv_dataframe(file_bytes: bytes) -> pd.DataFrame:
    encodings = ("utf-8-sig", "utf-8", "latin-1", "cp1252")
    separators = (";", ",", None)
    best_dataframe: pd.DataFrame | None = None
    best_column_count = 0
    last_error: Exception | None = None

    for encoding in encodings:
        for separator in separators:
            try:
                read_kwargs: dict[str, Any] = {"encoding": encoding}
                if separator is None:
                    read_kwargs["sep"] = None
                    read_kwargs["engine"] = "python"
                else:
                    read_kwargs["sep"] = separator

                dataframe = pd.read_csv(io.BytesIO(file_bytes), **read_kwargs)
                column_count = len(dataframe.columns)
                if column_count > best_column_count:
                    best_column_count = column_count
                    best_dataframe = dataframe
            except Exception as exc:
                last_error = exc

    if best_dataframe is not None and best_column_count > 0:
        return best_dataframe

    if last_error:
        raise ValueError(
            "Não foi possível ler o arquivo CSV. Verifique encoding e separador (; ou ,)."
        ) from last_error

    raise ValueError("Não foi possível ler o arquivo CSV.")


def _read_dataframe(file_bytes: bytes, filename: str) -> pd.DataFrame:
    lower_name = filename.lower()

    if lower_name.endswith(".csv"):
        return _read_csv_dataframe(file_bytes)

    if lower_name.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(file_bytes))

    raise ValueError("Formato não suportado. Envie um arquivo .xlsx, .xls ou .csv.")


def extract_addresses(file_bytes: bytes, filename: str) -> dict[str, Any]:
    dataframe = _read_dataframe(file_bytes, filename)
    column_map = _build_column_map(list(dataframe.columns))

    if not any(column_map.values()):
        raise ValueError(
            "Não encontrei colunas de endereço na planilha. "
            "Esperado algo como ENDERECO, NUMERO, BAIRRO, CEP, CIDADE e UF."
        )

    addresses: list[dict[str, Any]] = []
    completados_viacep = 0

    for index, row in dataframe.iterrows():
        formatted, enriched = _format_address(row, column_map)
        if not formatted:
            continue

        if enriched:
            completados_viacep += 1

        addresses.append(
            {
                "linha": int(index) + 2,
                "endereco": formatted,
            }
        )

    return {
        "arquivo": filename,
        "total_linhas_planilha": int(len(dataframe)),
        "total_enderecos": len(addresses),
        "enderecos_completados_viacep": completados_viacep,
        "enderecos": addresses,
        "colunas_detectadas": {
            key: column_map[key] for key in column_map if column_map[key]
        },
    }


def _result_to_row(item: dict[str, Any]) -> dict[str, Any]:
    distancia = item.get("distancia_km")
    return {
        "Linha": item.get("linha", ""),
        "Endereco": item.get("endereco", ""),
        "Viabilidade": item.get("viabilidade", ""),
        "Camada NIO": item.get("camada_nio", ""),
        "Distancia km": distancia if distancia is not None else "",
        "Status": item.get("status", ""),
        "Latitude": item.get("lat", ""),
        "Longitude": item.get("lng", ""),
        "Mensagem": item.get("mensagem", ""),
    }


def _sort_key(item: dict[str, Any]) -> tuple[int, float]:
    viabilidade = item.get("viabilidade", "")
    priority = {
        "Dentro da mancha": 0,
        "Próximo da mancha": 1,
    }.get(viabilidade, 2)
    distancia = item.get("distancia_km")
    distance_value = float(distancia) if distancia is not None else float("inf")
    return priority, distance_value


def _apply_row_highlight(worksheet, row_index: int, viabilidade: str, column_count: int) -> None:
    if viabilidade == "Dentro da mancha":
        fill = FILL_DENTRO
    elif viabilidade == "Próximo da mancha":
        fill = FILL_PROXIMO
    else:
        return

    for column in range(1, column_count + 1):
        cell = worksheet.cell(row=row_index, column=column)
        cell.fill = fill
        cell.font = FONT_DESTAQUE


def export_results_spreadsheet(
    resultados: list[dict[str, Any]],
    *,
    parcial: bool = False,
) -> bytes:
    viable_results = filter_mancha_verde_results(resultados)
    sorted_results = sorted(viable_results, key=_sort_key)
    rows = [_result_to_row(item) for item in sorted_results]
    column_names = [
        "Linha",
        "Endereco",
        "Viabilidade",
        "Camada NIO",
        "Distancia km",
        "Status",
        "Latitude",
        "Longitude",
        "Mensagem",
    ]

    summary_rows = [
        {"Informacao": "Total processado", "Valor": len(resultados)},
        {"Informacao": "Na mancha verde", "Valor": len(viable_results)},
        {
            "Informacao": "Observacao",
            "Valor": (
                "Exportação parcial: nem todos os endereços foram processados."
                if parcial and viable_results
                else "Nenhum endereço ficou dentro da mancha verde."
                if not viable_results
                else "Somente endereços dentro da mancha verde são listados abaixo."
            ),
        },
    ]

    legend_rows = [
        {"Coluna": "Linha", "Significado": "Número da linha na planilha original"},
        {"Coluna": "Endereco", "Significado": "Texto que foi buscado no My Maps"},
        {
            "Coluna": "Viabilidade",
            "Significado": "Somente endereços dentro da mancha verde são exportados",
        },
        {
            "Coluna": "Camada NIO",
            "Significado": "Camada NIO onde o endereço caiu",
        },
        {
            "Coluna": "Distancia km",
            "Significado": "0 significa dentro da mancha verde",
        },
        {
            "Coluna": "Status",
            "Significado": "Se a automação conseguiu processar aquele endereço",
        },
        {
            "Coluna": "Latitude",
            "Significado": "Coordenada usada para verificar a camada (norte/sul)",
        },
        {
            "Coluna": "Longitude",
            "Significado": "Coordenada usada para verificar a camada (leste/oeste)",
        },
        {"Coluna": "Mensagem", "Significado": "Detalhe do erro, se houver"},
        {
            "Coluna": "Destaque visual",
            "Significado": "Verde = endereço dentro da mancha verde",
        },
    ]

    results_df = pd.DataFrame(rows, columns=column_names)
    summary_df = pd.DataFrame(summary_rows)
    legend_df = pd.DataFrame(legend_rows)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Resumo")
        sheet_name = "Enderecos na mancha verde"
        results_df.to_excel(writer, index=False, sheet_name=sheet_name)
        legend_df.to_excel(writer, index=False, sheet_name="Legenda das colunas")

        if not results_df.empty:
            worksheet = writer.sheets[sheet_name]
            column_count = len(results_df.columns)
            for row_index, item in enumerate(sorted_results, start=2):
                _apply_row_highlight(
                    worksheet,
                    row_index,
                    str(item.get("viabilidade", "")),
                    column_count,
                )

    return buffer.getvalue()
